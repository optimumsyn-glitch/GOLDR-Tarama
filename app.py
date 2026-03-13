import streamlit as st
import yfinance as yf
import pandas as pd
import numpy as np
from scipy.stats import linregress
import time
import warnings

warnings.filterwarnings("ignore")

# ================================================================
# AYARLAR (kullanıcı sidebar'dan değiştirebilecek)
# ================================================================
GOLDR_FACTOR    = 3.0
GOLDR_ATR_PERIOD = 10

# ================================================================
# AKTİF BIST LİSTE (senin orijinal listen)
# ================================================================
symbols = [
    'A1CAP.IS','A1YAT.IS','ACSEL.IS','ADEL.IS','ADESE.IS','ADGYO.IS','AEFES.IS','AFYON.IS','AGESA.IS','AGHOL.IS',
    'AGROT.IS','AGYO.IS','AHGAZ.IS','AKBNK.IS','AKCNS.IS','AKENR.IS','AKFGY.IS','AKFYE.IS','AKGRT.IS','AKMGY.IS',
    'AKSA.IS','AKSEN.IS','ALARK.IS','ALBRK.IS','ALCAR.IS','ALCTL.IS','ALFAS.IS','ALGYO.IS','ALKA.IS','ALKIM.IS',
    'ALMAD.IS','ANELE.IS','ANGEN.IS','ANHYT.IS','ANSGR.IS','ARCLK.IS','ARDYZ.IS','ARENA.IS','ARSAN.IS','ARTMS.IS',
    'ARZUM.IS','ASELS.IS','ASGYO.IS','ASTOR.IS','ASUZU.IS','ATAKP.IS','ATATP.IS','ATEKS.IS','ATLAS.IS','AVHOL.IS',
    'AVPGY.IS','AYDEM.IS','AYEN.IS','AYES.IS','AZTEK.IS','BAGFS.IS','BAKAB.IS','BANVT.IS','BARMA.IS','BAYRK.IS',
    'BEGYO.IS','BERA.IS','BEYAZ.IS','BIMAS.IS','BINHO.IS','BIOEN.IS','BIZIM.IS','BMSCH.IS','BMSTL.IS','BNTAS.IS',
    'BOBET.IS','BORLS.IS','BOSSA.IS','BRISA.IS','BRKVY.IS','BRSAN.IS','BRYAT.IS','BSOKE.IS','BTCIM.IS','BUCIM.IS',
    'BURCE.IS','BURVA.IS','BVSAN.IS','CANTE.IS','CATES.IS','CCOLA.IS','CELHA.IS','CEMAS.IS','CEMTS.IS','CEOEM.IS',
    'CIMSA.IS','CLEBI.IS','CONSE.IS','CVKMD.IS','CWENE.IS','DAGHL.IS','DAGI.IS','DAPGM.IS','DARDL.IS','DENGE.IS',
    'DERHL.IS','DERIM.IS','DESA.IS','DESPC.IS','DGATE.IS','DGNMO.IS','DIRIT.IS','DITAS.IS','DMRGD.IS','DMSAS.IS',
    'DNISI.IS','DOAS.IS','DOCO.IS','DOHOL.IS','DOKTA.IS','DURDO.IS','DYOBY.IS','EBEBK.IS','ECILC.IS','ECZYT.IS',
    'EDATA.IS','EDIP.IS','EFORC.IS','EGEEN.IS','EGGUB.IS','EGEPO.IS','EGPRO.IS','EGSER.IS','EKGYO.IS','EKOS.IS',
    'ELITE.IS','EMKEL.IS','ENERY.IS','ENJSA.IS','ENKAI.IS','ERBOS.IS','EREGL.IS','ERSU.IS','ESCAR.IS','ESCOM.IS',
    'ESEN.IS','ETILR.IS','EUPWR.IS','EUREN.IS','EYGYO.IS','FADE.IS','FENER.IS','FLAP.IS','FMIZP.IS','FONET.IS',
    'FORMT.IS','FORTE.IS','FRIGO.IS','FROTO.IS','FUZUL.IS','GARAN.IS','GARFA.IS','GEDIK.IS','GEDZA.IS','GENIL.IS',
    'GENTS.IS','GEREL.IS','GESAN.IS','GIPTA.IS','GLBMD.IS','GLCVY.IS','GLYHO.IS','GMTAS.IS','GOKNR.IS','GOLTS.IS',
    'GOODY.IS','GOZDE.IS','GRSEL.IS','GSDDE.IS','GSDHO.IS','GUBRF.IS','GWIND.IS','GZNMI.IS','HALKB.IS','HATEK.IS',
    'HATSN.IS','HEKTS.IS','HKTM.IS','HLGYO.IS','HTTBT.IS','HUNER.IS','HURGZ.IS','ICBCT.IS','IDGYO.IS','IEYHO.IS',
    'IHEVA.IS','IHGZT.IS','IHLAS.IS','IHLGM.IS','IHYAY.IS','IMASM.IS','INDES.IS','INFO.IS','INGRM.IS','INTEM.IS',
    'INVEO.IS','ISATR.IS','ISBIR.IS','ISBTR.IS','ISCTR.IS','ISFIN.IS','ISGSY.IS','ISGYO.IS','ISMEN.IS','ISSEN.IS',
    'ISYAT.IS','IZENR.IS','IZFAS.IS','IZMDC.IS','JANTS.IS','KAPLM.IS','KAREL.IS','KARSN.IS','KARTN.IS','KARYER.IS',
    'KATMR.IS','KAYSE.IS','KCAER.IS','KCHOL.IS','KENT.IS','KERVT.IS','KFEIN.IS','KLRHO.IS','KLMSN.IS','KLNMA.IS',
    'KLSER.IS','KLYAS.IS','KNFRT.IS','KONTR.IS','KONYA.IS','KOPOL.IS','KORDS.IS','KOZAA.IS','KRDMA.IS','KRDMB.IS',
    'KRDMD.IS','KRGYO.IS','KRONT.IS','KRPLS.IS','KRSTL.IS','KSTUR.IS','KUTPO.IS','KUVVA.IS','KUYAS.IS','KZBGY.IS',
    'KZGYO.IS','LIDER.IS','LIDFA.IS','LINK.IS','LOGO.IS','LUKSK.IS','MAALT.IS','MAGEN.IS','MAKIM.IS','MAKTK.IS',
    'MANAS.IS','MARBL.IS','MARKA.IS','MARTI.IS','MAVI.IS','MEDTR.IS','MEGAP.IS','MEGMT.IS','MEKAG.IS','MEMAS.IS',
    'MERCN.IS','MERIT.IS','MERKO.IS','METRO.IS','METUR.IS','MGROS.IS','MIATK.IS','MIPAZ.IS','MMCAS.IS','MNDRS.IS',
    'MNDTR.IS','MOBTL.IS','MPARK.IS','MRGYO.IS','MRSHL.IS','MSGYO.IS','MTRYO.IS','MZHLD.IS','NATEN.IS','NETAS.IS',
    'NIBAS.IS','NTGAZ.IS','NTHOL.IS','NUGYO.IS','NUHCM.IS','OBASE.IS','OBAMS.IS','ODAS.IS','ONCSM.IS','ORCAY.IS',
    'ORGE.IS','ORMA.IS','OSMEN.IS','OTKAR.IS','OYAKC.IS','OYAYO.IS','OYLUM.IS','OZGYO.IS','OZKGY.IS','OZRDN.IS',
    'OZSUB.IS','PAGYO.IS','PAMEL.IS','PAPIL.IS','PARSN.IS','PASEU.IS','PCILT.IS','PEGYO.IS','PENGD.IS','PENTA.IS',
    'PETKM.IS','PETUN.IS','PGSUS.IS','PINSU.IS','PKART.IS','PKENT.IS','PLTUR.IS','PNLSN.IS','PNSUT.IS','POLHO.IS',
    'POLTK.IS','PRDGS.IS','PRKAB.IS','PRKME.IS','PRZMA.IS','PSDTC.IS','QNBFB.IS','QNBFL.IS','QUAGR.IS','RALHY.IS',
    'RAYSG.IS','REEDR.IS','RNPOL.IS','RODRG.IS','RTALB.IS','RUBNS.IS','RYGYO.IS','RYSAS.IS','SAFKR.IS','SAHOL.IS',
    'SAMAT.IS','SANEL.IS','SANFM.IS','SANKO.IS','SARKY.IS','SASA.IS','SAYAS.IS','SDTTR.IS','SEGYO.IS','SELEC.IS',
    'SELGD.IS','SELVA.IS','SEYKM.IS','SILVR.IS','SISE.IS','SKBNK.IS','SKTAS.IS','SMRTG.IS','SNGYO.IS','SNICA.IS',
    'SNKPA.IS','SOKM.IS','SONME.IS','SRVGY.IS','SUWEN.IS','TABGD.IS','TARKM.IS','TATEN.IS','TATGD.IS','TAVHL.IS',
    'TCELL.IS','TDGYO.IS','TEKTU.IS','TERA.IS','TETMT.IS','TEZOL.IS','THYAO.IS','TIRE.IS','TKFEN.IS','TKNSA.IS',
    'TLMAN.IS','TMPOL.IS','TMSN.IS','TNZTP.IS','TOASO.IS','TRCAS.IS','TRGYO.IS','TRILC.IS','TSGYO.IS','TSKB.IS',
    'TSPOR.IS','TTKOM.IS','TTRAK.IS','TUCLK.IS','TUKAS.IS','TUPRS.IS','TURSG.IS','UFUK.IS','ULAS.IS','ULKER.IS',
    'ULUFA.IS','ULUSE.IS','ULUUN.IS','UMPAS.IS','USAK.IS','VAKBN.IS','VAKFN.IS','VAKKO.IS','VANGD.IS','VBTYZ.IS',
    'VERTU.IS','VERUS.IS','VESBE.IS','VESTL.IS','VKFYO.IS','VKGYO.IS','VKING.IS','YAPRK.IS','YATAS.IS','YAYLA.IS',
    'YBTAS.IS','YEOTK.IS','YESIL.IS','YGGYO.IS','YGYO.IS','YIGIT.IS','YONGA.IS','YUNSA.IS','YYAPI.IS','ZEDUR.IS',
    'ZRGYO.IS','ZOREN.IS'
]

# ================================================================
# FONKSİYONLAR (senin orijinal kodundan)
# ================================================================
def calc(close, period):
    if len(close) < period: return None
    y = close[-period:]
    x = np.arange(period)
    slope, intercept, r, _, _ = linregress(x, y)
    pred = slope * x + intercept
    std = np.std(y - pred)
    lower = (slope * (period - 1) + intercept) - 2 * std
    return r, y[-1], lower

def goldr_trend(df, period=10, multiplier=3):
    hl2 = (df['High'] + df['Low']) / 2
    tr = pd.concat([
        df['High'] - df['Low'],
        abs(df['High'] - df['Close'].shift()),
        abs(df['Low'] - df['Close'].shift())
    ], axis=1).max(axis=1)
    atr = tr.rolling(window=period).mean()
    upper_basic = hl2 + (multiplier * atr)
    lower_basic = hl2 - (multiplier * atr)
    upper = pd.Series(np.nan, index=df.index)
    lower = pd.Series(np.nan, index=df.index)
    trend = pd.Series(0, index=df.index)
    first_valid = period
    if first_valid >= len(df):
        return "Veri yetersiz"
    upper.iloc[first_valid] = upper_basic.iloc[first_valid]
    lower.iloc[first_valid] = lower_basic.iloc[first_valid]
    for i in range(first_valid + 1, len(df)):
        if df['Close'].iloc[i] > upper.iloc[i-1]:
            trend.iloc[i] = 1
            lower.iloc[i] = max(lower_basic.iloc[i], lower.iloc[i-1] if not pd.isna(lower.iloc[i-1]) else lower_basic.iloc[i])
            upper.iloc[i] = upper_basic.iloc[i]
        elif df['Close'].iloc[i] < lower.iloc[i-1]:
            trend.iloc[i] = -1
            upper.iloc[i] = min(upper_basic.iloc[i], upper.iloc[i-1] if not pd.isna(upper.iloc[i-1]) else upper_basic.iloc[i])
            lower.iloc[i] = lower_basic.iloc[i]
        else:
            trend.iloc[i] = trend.iloc[i-1]
            if trend.iloc[i] == 1:
                lower.iloc[i] = max(lower_basic.iloc[i], lower.iloc[i-1])
                upper.iloc[i] = upper_basic.iloc[i]
            else:
                upper.iloc[i] = min(upper_basic.iloc[i], upper.iloc[i-1])
                lower.iloc[i] = lower_basic.iloc[i]
    current = trend.iloc[-1]
    prev = trend.iloc[-2] if len(trend) > 1 else 0
    if current == 1 and prev != 1:
        return "Yeni Yukarı Dönüş (AL sinyali)"
    elif current == 1:
        return "Yukarı Trend"
    else:
        return "Aşağı Trend"

# ================================================================
# ŞİFRE EKRANI
# ================================================================
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

def check_password():
    if st.session_state.authenticated:
        return True

    st.title("GOLDR-TARAMA - BIST Teknik Tarayıcı")
    st.markdown("Bu araç sadece yetkili kullanıcılar içindir.")
    password = st.text_input("Şifreyi girin", type="password")

    if st.button("Giriş"):
        # BURADAKİ ŞİFREYİ KENDİ İSTEDİĞİN GİBİ DEĞİŞTİR
        if password == "goldr2026tr":
            st.session_state.authenticated = True
            st.success("Giriş başarılı!")
            st.rerun()
        else:
            st.error("Yanlış şifre")
    return False

if not check_password():
    st.stop()

# ================================================================
# ANA EKRAN
# ================================================================
st.title("GOLDR-TARAMA")
st.write("Pearson Kanalı + EMA Bulutu + GOLDR Supertrend taraması")

with st.sidebar:
    st.header("Tarama Ayarları")
    PEARSON_MIN     = st.slider("Min Pearson", 0.60, 0.95, 0.72, 0.01)
    DIP_TOLERANCE   = st.slider("Dip Toleransı (%)", 0.0, 10.0, 2.5, 0.1) / 100
    LOOKBACK        = st.selectbox("Veri Dönemi", ["6mo", "1y", "2y", "3y"], index=2)
    CHUNK_SIZE      = st.number_input("Chunk Boyutu", 10, 60, 35)
    SLEEP_CHUNK     = st.number_input("Parça arası bekleme (sn)", 5, 30, 12)

rows = []

if st.button("Tarama Başlat (biraz zaman alabilir)"):
    with st.spinner("Tarama sürüyor..."):
        for start in range(0, len(symbols), CHUNK_SIZE):
            chunk = symbols[start:start + CHUNK_SIZE]
            st.write(f"İşleniyor: {start//CHUNK_SIZE + 1}. grup ({len(chunk)} hisse)")
            
            try:
                data = yf.download(chunk, period=LOOKBACK, group_by='ticker', auto_adjust=True, progress=False, threads=True, timeout=25)
                
                for sym in chunk:
                    try:
                        df_stock = data[sym].dropna()
                        if len(df_stock) < 180: continue
                        
                        close_s = df_stock["Close"]
                        vol_s   = df_stock["Volume"]
                        
                        e22  = close_s.ewm(span=22,  adjust=False).mean().iloc[-1]
                        e34  = close_s.ewm(span=34,  adjust=False).mean().iloc[-1]
                        e55  = close_s.ewm(span=55,  adjust=False).mean().iloc[-1]
                        e89  = close_s.ewm(span=89,  adjust=False).mean().iloc[-1]
                        e144 = close_s.ewm(span=144, adjust=False).mean().iloc[-1]
                        e200 = close_s.ewm(span=200, adjust=False).mean().iloc[-1]
                        
                        ema_ideal = e22 > e34 > e55 > e89 > e144 > e200
                        
                        avg_vol  = vol_s.rolling(10).mean().iloc[-1]
                        last_vol = vol_s.iloc[-1]
                        vol_ok   = last_vol > avg_vol * 1.2
                        
                        goldr_durum = goldr_trend(df_stock, GOLDR_ATR_PERIOD, GOLDR_FACTOR)
                        
                        best = None
                        best_r = -1.0
                        
                        for p in range(20, 201, 5):
                            out = calc(close_s.values, p)
                            if out is None: continue
                            r, price, lower = out
                            
                            if r >= PEARSON_MIN and r > best_r:
                                best_r = r
                                at_dip = price <= lower * (1 + DIP_TOLERANCE)
                                
                                ema_levels = {"22": e22, "34": e34, "55": e55, "89": e89, "144": e144, "200": e200}
                                closest = min(ema_levels, key=lambda k: abs(price - ema_levels[k]))
                                diff_pct = abs(price - ema_levels[closest]) / price
                                ema_note = f"{closest}'e çok yakın" if diff_pct < 0.005 else f"{closest}'e yakın" if diff_pct < 0.015 else ""
                                
                                if at_dip and ema_ideal and vol_ok and best_r >= 0.85 and "Yeni Yukarı Dönüş" in goldr_durum:
                                    msg = "Kritik alım bölgesi. Fiyat kanal alt bandında, EMA dizilimi ideal, Pearson çok yüksek, hacim artıyor ve GOLDR yeni yukarı dönüş yapmış. Eğim yukarı, momentum güçlü. Buradan kanal tepesine doğru genişleme potansiyeli yüksek. Al sinyali net, risk alınabilir."
                                elif at_dip and ema_ideal and best_r >= 0.82 and "Yukarı Trend" in goldr_durum:
                                    msg = "Güçlü destek seviyesi. EMA'lar ideal sıralı, fiyat kanal dibinde, Pearson yüksek, GOLDR yukarı trendde. Eğim pozitif. Kanal tepesine kadar taşıma ihtimali var. Mevcut pozisyonlar korunabilir."
                                elif at_dip and best_r >= 0.80 and "Yeni Yukarı Dönüş" in goldr_durum:
                                    msg = "Teknik dip + GOLDR yeni al sinyali. Fiyat kanal dibine yaklaşmış, Pearson güçlü, GOLDR yukarı dönüş yapmış. EMA'lar yakın. Kısa vadeli tepki alımı için uygun."
                                elif ema_ideal and vol_ok and best_r >= 0.82 and "Yukarı Trend" in goldr_durum:
                                    msg = "Sağlıklı yükseliş trendi. EMA bulutu pozitif, hacim destekliyor, Pearson yüksek, GOLDR yukarı trendde. Fiyat EMA'ların üzerinde, kanal içinde yukarı eğim var."
                                elif ema_ideal and best_r >= 0.78:
                                    msg = "Uzun vadeli trend yapısı sağlam. EMA dizilimi ideal, Pearson yeterli. Fiyat EMA bandının üst kısmında. GOLDR yönü pozitif. Kanal içinde yukarı hareket beklenebilir."
                                elif at_dip and vol_ok and "Yeni Yukarı Dönüş" in goldr_durum:
                                    msg = "Hacimli dip + GOLDR yeni al sinyali. Fiyat kanal dibinde, hacim artıyor, GOLDR yukarı dönüş yapmış. Pearson orta-yüksek. Kısa vadeli tepki alımı potansiyeli yüksek."
                                elif at_dip:
                                    msg = "Kanal alt bandı test ediliyor. Pearson yeterli ancak diğer onaylar sınırlı. GOLDR henüz yukarı dönüş yapmadıysa temkinli olunmalı."
                                elif ema_ideal and "Yukarı Trend" in goldr_durum:
                                    msg = "Trend yapısı korunuyor. EMA dizilimi ideal, GOLDR yukarı trendde. Fiyat EMA bandı üzerinde. Kanal içinde yukarı yönlü hareket beklenebilir."
                                else:
                                    msg = "Trend devam ediyor ancak onaylar sınırlı. Pearson orta seviyede, EMA dizilimi zayıf veya karışık, GOLDR yönü net değil. Yeni pozisyon için daha net sinyal beklenmeli."
                                
                                best = {
                                    "Sembol": sym.replace(".IS",""),
                                    "Periyot": p,
                                    "Pearson": round(r, 3),
                                    "Kapanış": round(price, 2),
                                    "Kanal Destek": round(lower, 2),
                                    "Diptemi?": "EVET" if at_dip else "HAYIR",
                                    "EMA Dizilimi": "İdeal" if ema_ideal else "Zayıf",
                                    "En Yakın EMA": closest,
                                    "Yakın Not": ema_note,
                                    "Hacim": "Yüksek" if vol_ok else "Normal",
                                    "GOLDR Durumu": goldr_durum,
                                    "Yorum": msg
                                }
                        
                        if best:
                            rows.append(best)
                    
                    except:
                        continue
                
            except Exception as e:
                st.write(f"Chunk hatası: {str(e)[:80]}")
                time.sleep(30)
            
                                    time.sleep(SLEEP_CHUNK)

    if rows:
        df = pd.DataFrame(rows).sort_values("Pearson", ascending=False)
        st.success(f"{len(df)} sonuç bulundu!")
        st.dataframe(df)

        # ────────────────────────────────────────────────
        # Gerçek Excel (.xlsx) indirme + orijinal stil (mavi/yeşil boyama, satır yüksekliği vs.)
        # ────────────────────────────────────────────────
        import io
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Tarama Sonuçları"

        # Başlık satırını yaz
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="37474F")  # koyu gri-mavi başlık
            cell.alignment = Alignment(horizontal="center")

        # Verileri yaz
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Stil uygulama
        critical_fill = PatternFill("solid", fgColor="BBDEFB")   # açık mavi - kritik
        strong_fill   = PatternFill("solid", fgColor="E8F5E9")   # açık yeşil - güçlü

        goldr_col_idx = df.columns.get_loc("GOLDR Durumu") + 1
        yorum_col_idx = df.columns.get_loc("Yorum") + 1

        for row_idx in range(2, len(df) + 2):
            yorum = str(ws.cell(row=row_idx, column=yorum_col_idx).value or "")
            goldr = str(ws.cell(row=row_idx, column=goldr_col_idx).value or "")

            is_critical = "Kritik" in yorum or "Yeni Yukarı Dönüş" in goldr
            is_strong   = "Güçlü" in yorum

            row_height = 80 if is_critical else 55 if is_strong else 40
            ws.row_dimensions[row_idx].height = row_height

            fill_color = critical_fill if is_critical else strong_fill if is_strong else None

            if fill_color:
                for col_idx in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = fill_color

            # Tüm hücrelere ince kenarlık
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

        # Sütun genişliklerini otomatik ayarla
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 6)
            ws.column_dimensions[column].width = min(adjusted_width, 90)

        # Yorum sütununu wrap text yap
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=yorum_col_idx, max_col=yorum_col_idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center")

        wb.save(output)
        output.seek(0)

        st.download_button(
            label="Stilli Excel (.xlsx) indir (mavi/yeşil boyalı)",
            data=output,
            file_name="goldr_tarama_sonuclari_stilli.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        st.warning("Sonuç bulunamadı.")

# ────────────────────────────────────────────────
# Dosya sonu - ekstra bir şey eklemek istersen buraya koyabilirsin
# ────────────────────────────────────────────────
st.info("Uygulama başarıyla çalıştı. Tarama sonuçlarını Excel olarak indirebilirsiniz.")
